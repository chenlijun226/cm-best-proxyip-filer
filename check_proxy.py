import requests
import pandas as pd
import os
import threading
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from concurrent.futures import ThreadPoolExecutor, as_completed

# --- 基础配置 ---
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
IP_FILE = os.path.join(CURRENT_DIR, 'ip.txt')
EXCEL_FILE = os.path.join(CURRENT_DIR, 'result.xlsx')
TXT_SUCCESS_FILE = os.path.join(CURRENT_DIR, 'success.txt')

# 线程数
MAX_WORKERS = 30
# 线程锁
file_lock = threading.Lock()

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36'
}

COLUMNS = ['success', 'proxyIP', 'portRemote', 'colo', 'responseTime', 'message', 'timeStamp']


def read_ips(file_path):
    if not os.path.exists(file_path):
        print(f"错误：找不到文件 {file_path}")
        return []
    pairs = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if ':' in line:
                parts = line.split(':')
                if len(parts) >= 2:
                    pairs.append((parts[0].strip(), parts[1].strip()))
    return pairs


def save_to_success_txt(data):
    """提取指定信息并追加到 success.txt"""
    try:
        probe = data.get('probe_results', {})
        # 优先取 ipv4，若无则取 ipv6 (针对 ipv6_only 的代理)
        target = probe.get('ipv4') if probe.get('ipv4') else probe.get('ipv6')

        if not target:
            return

        candidate = target.get('candidate', 'N/A')
        exit_info = target.get('exit', {})
        country = exit_info.get('country', '')
        city = exit_info.get('city', '')
        org = exit_info.get('org', '')

        # 格式：23.105.206.238:9443#US Los Angeles AS25820 IT7 Networks Inc
        line = f"{candidate}#{country} {city} {org}".strip()

        with file_lock:
            with open(TXT_SUCCESS_FILE, 'a', encoding='utf-8') as f:
                f.write(line + '\n')
    except Exception as e:
        print(f"写入 TXT 失败: {e}")


def check_proxy_task(ip, port):
    url = f'https://api.090227.xyz/check?proxyip={ip}:{port}'
    no_proxies = {"http": None, "https": None}

    try:
        response = requests.get(url, headers=HEADERS, timeout=15, proxies=no_proxies, verify=True)
        if response.status_code == 200:
            result = response.json()
            success_val = result.get('success')
            if success_val is True or str(success_val).lower() == 'true':
                save_to_success_txt(result)
                return result
    except Exception:
        pass
    return None


def save_to_excel(new_data_list):
    if not new_data_list:
        return

    if os.path.exists(EXCEL_FILE):
        try:
            old_df = pd.read_excel(EXCEL_FILE)
        except:
            old_df = pd.DataFrame(columns=COLUMNS)
    else:
        old_df = pd.DataFrame(columns=COLUMNS)

    new_df = pd.DataFrame(new_data_list).reindex(columns=COLUMNS)
    final_df = pd.concat([old_df, new_df], ignore_index=True)

    if not final_df.empty:
        final_df['proxyIP'] = final_df['proxyIP'].astype(str).str.strip()
        final_df['portRemote'] = final_df['portRemote'].astype(str).str.strip()
        final_df = final_df.drop_duplicates(subset=['proxyIP', 'portRemote'], keep='last')

    final_df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')

    # 优化时间显示
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        style = NamedStyle(name="ts_style", number_format="yyyy-mm-dd hh:mm:ss")
        col_idx = COLUMNS.index('timeStamp') + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                try:
                    val_str = str(cell.value).replace('Z', '').split('.')[0]
                    if 'T' in val_str:
                        cell.value = datetime.fromisoformat(val_str)
                        cell.style = style
                except:
                    pass
        wb.save(EXCEL_FILE)
    except:
        pass


def main():
    print(f"开始多线程检测 (线程数: {MAX_WORKERS})...")
    ips = read_ips(IP_FILE)
    if not ips:
        print("没有可检测的数据。")
        return

    valid_results = []
    total = len(ips)
    completed = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_proxy = {executor.submit(check_proxy_task, ip, port): (ip, port) for ip, port in ips}

        for future in as_completed(future_to_proxy):
            completed += 1
            ip, port = future_to_proxy[future]
            try:
                data = future.result()
                if data:
                    valid_results.append(data)
                    print(f"[{completed}/{total}] {ip}:{port} -> 【可用】")
                else:
                    print(f"[{completed}/{total}] {ip}:{port} -> 【不可用】")
            except Exception as e:
                print(f"[{completed}/{total}] {ip}:{port} -> 【错误: {e}】")

    # 保存并打印你要求的统计信息
    count = len(valid_results)
    if count > 0:
        save_to_excel(valid_results)
        print(f"\n检测完成！Excel 总记录已更新{count}个，success.txt 已追加有效代理{count}个。")
    else:
        print("\n检测完成，本次未发现新的有效代理。")


if __name__ == '__main__':
    main()